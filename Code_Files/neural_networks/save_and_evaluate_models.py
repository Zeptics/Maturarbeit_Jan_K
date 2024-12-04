import os
import subprocess
import time
import tensorflow as tf
import numpy as np
from tqdm import tqdm
import xml.etree.ElementTree as ET
import cv2
import openpyxl
from openpyxl import styles
import ast


'''
************************************************************************************************************************
GENERAL DEFINITIONS:
************************************************************************************************************************
'''


def get_max_column_with_data(sheet):
    max_column = 0
    # Loop through all rows to check each row's actual max column with data
    for row in sheet.iter_rows():
        for cell in reversed(row):  # Start checking from the last column
            if cell.value is not None:
                max_column = max(max_column, cell.column)
                break  # Stop at the first non-empty cell in the row
    return max_column


def get_max_row_with_data(sheet):
    max_row = 0
    # Loop through all rows to check each row's actual max column with data
    for column in sheet.iter_cols():
        for cell in reversed(column):  # Start checking from the last column
            if cell.value is not None:
                max_row = max(max_row, cell.row)
                break  # Stop at the first non-empty cell in the row
    return max_row


'''
************************************************************************************************************************
STEP 1: SAVING THE MODELS FROM CHECKPOINT FILES (--> saved_model.pb file)
************************************************************************************************************************
'''


def create_checkpoint_file(parent_dir, step_n_dir, all_checkpoints, training_steps):
    # if it's the final checkpoint folder add 1 to the name as the there is an extra ckpt file for the final ckpt (with +1)
    if int(step_n_dir[4:]) == training_steps:
        checkpoint_content = f'model_checkpoint_path: "ckpt-{int((int(step_n_dir[4:]) /100) + 1)}"\n'
    else:
        print(int(int(step_n_dir[4:])/100))
        checkpoint_content = f'model_checkpoint_path: "ckpt-{int(int(step_n_dir[4:])/100)}"\n'

    print(checkpoint_content)
    for checkpoint in all_checkpoints:
        checkpoint_content += f'all_model_checkpoint_paths: "{checkpoint}"\n'

    # Write the content to a file named "checkpoint"
    checkpoint_file_path = os.path.join(parent_dir, step_n_dir, 'checkpoint')
    os.makedirs(os.path.dirname(checkpoint_file_path), exist_ok=True)
    print(checkpoint_file_path)
    # with open(checkpoint_file_path, 'w') as f:
    #     f.write(checkpoint_content)

    checkpoint_file = open(checkpoint_file_path, 'w')

    checkpoint_file.write(checkpoint_content)

    checkpoint_file.close()

    print(f'Checkpoint file created at: {checkpoint_file_path}')


def run_script_in_conda_env(conda_env_name, target_dir, script_name, script_flags):
    # Construct the command to run in conda env
    command = f'cmd /c "cd /d {target_dir} && conda activate {conda_env_name} && python {script_name} {script_flags}"'

    print(command)

    # Run the command
    process = subprocess.run(command, shell=True)

    if process.returncode != 0:
        print(f'Command failed with return code {process.returncode}')


def save_models(parent_dir, all_checkpoints, conda_env_name, root_dir, script_name, model_name, model_version, training_steps):
    for step_dir in [d for d in os.listdir(parent_dir) if d not in already_treated_steps]:
        # Create a checkpoint file
        create_checkpoint_file(parent_dir, step_dir, all_checkpoints, training_steps)
        # Make "saved_model" directory in the folder if it doesn't exist already
        os.makedirs(os.path.join(parent_dir, step_dir, 'saved_model'), exist_ok=True)
    for step_dir in [d for d in os.listdir(parent_dir) if d not in already_treated_steps]:
        print(f'saving {step_dir}')
        # Define script flags
        script_flags = "--input_type=image_tensor " \
                       f"--pipeline_config_path={os.path.join(os.getcwd(), 'Tensorflow', 'workspace', 'models', model_name, model_version, 'pipeline.config')} " \
                       f"--trained_checkpoint_dir={os.path.join(os.getcwd(), 'Tensorflow', 'workspace', 'models', model_name, model_version, 'checkpoints_steps', step_dir)} " \
                       f"--output_directory={os.path.join(os.getcwd(), 'Tensorflow', 'workspace', 'models', model_name, model_version, 'checkpoints_steps', step_dir, 'saved_model')} " \
            # Save the model
        run_script_in_conda_env(conda_env_name, root_dir, script_name, script_flags)

'''
************************************************************************************************************************
STEP 2: PREDICT BOUNDING BOXES USING THE SAVED MODELS AND SAVE BBOXES TO XML FILES
************************************************************************************************************************
'''


def load_image_into_numpy_array(path):
    img_data = cv2.imread(path)
    return np.array(img_data)


def create_xml_files_for_detected_bboxes(image_name, xml_file_parent_dir, detections, scores, width, height, threshold_list):
    for threshold in threshold_list:
        # Initialize a new root for every xml file
        root = ET.Element('annotation')
        filename = ET.SubElement(root, 'filename')
        filename.text = image_name
        # Create object elements for each bounding box
        for idx, (y1, x1, y2, x2) in enumerate(detections, start=1):
            if scores[idx - 1] >= threshold:
                obj = ET.SubElement(root, 'object')
                obj_name = ET.SubElement(obj, 'name')
                obj_name.text = f'object_{idx}'
                bbox = ET.SubElement(obj, 'bndbox')
                xmin = ET.SubElement(bbox, 'xmin')
                xmin.text = str(int(width*x1))
                ymin = ET.SubElement(bbox, 'ymin')
                ymin.text = str(int(height*y1))
                xmax = ET.SubElement(bbox, 'xmax')
                xmax.text = str(int(width*x2))
                ymax = ET.SubElement(bbox, 'ymax')
                ymax.text = str(int(height*y2))

        # Create XML tree and write to file
        tree = ET.ElementTree(root)
        xml_filename = os.path.splitext(image_name)[0] + ".xml"
        xml_path = os.path.join(xml_file_parent_dir, f'threshold_{threshold}', xml_filename)
        tree.write(xml_path)


def add_parameters_to_xlsx(xlsx_file_path, avg_time_per_img, model_name, step_folder, warmup_steps):
    wb_obj = openpyxl.load_workbook(xlsx_file_path)

    sheet = wb_obj.active

    row = get_max_row_with_data(sheet)

    for i, threshold in enumerate(threshold_list):
        sheet[f'A{row + i + 1}'].value = f'{model_name}_{step_folder}'
        sheet[f'B{row + i + 1}'].value = step_folder[4:]
        sheet[f'C{row + i + 1}'].value = warmup_steps
        sheet[f'D{row + i + 1}'].value = threshold
        sheet[f'E{row + i + 1}'].value = avg_time_per_img

    wb_obj.save(xlsx_file_path)


def predict_bounding_boxes_with_models(parent_dir, images_dir, threshold_list):
    # Sorting them by amount of training steps rather than alphabetically
    steps_list = []
    for step_folder in [sf for sf in os.listdir(parent_dir) if sf not in already_treated_steps]:
        steps_list.append(int(step_folder[4:]))
    steps_list.sort()
    for i, step in enumerate(steps_list):
        steps_list[i] = f'step{step}'

    # Actually get predictions for every model now
    for step_folder in tqdm(steps_list):
        model_dir = os.path.join(parent_dir, step_folder, 'saved_model', 'saved_model')
        loaded_model = tf.saved_model.load(model_dir)

        # initialize start time to be able to calculate average time per image
        start_time = time.time()
        for image in [file for file in os.listdir(images_dir) if file.endswith('.png')]:
            image_path = os.path.join(images_dir, image)
            # convert png to numpy array
            image_np = load_image_into_numpy_array(image_path)
            # convert numpy array to tensorflow tensor (format in which tf takes data)
            input_tensor = tf.convert_to_tensor(image_np)
            # Add extra dimension to tensor so that it's compatible with the model
            input_tensor = input_tensor[tf.newaxis, ...]

            # Get width and height of image
            img_width, img_height = image_np.shape[:2]

            # Actually run the model and get the predictions
            detections = loaded_model(input_tensor)
            # 'detections' is a dictionary containing detection results
            # For example, 'detections['detection_boxes']' contains bounding box coordinates
            boxes = detections['detection_boxes'][0].numpy()
            classes = detections['detection_classes'][0].numpy().astype(np.int64)
            scores = detections['detection_scores'][0].numpy()

            # SAVE PREDICTIONS TO XML FILES
            # construct the xml output directory
            xml_out_dir = os.path.join(parent_dir, step_folder, 'predictions')
            # create all directories for the xml output files
            for threshold in threshold_list:
                os.makedirs(os.path.join(xml_out_dir, f'threshold_{threshold}'), exist_ok=True)
            create_xml_files_for_detected_bboxes(image, xml_out_dir, boxes, scores, img_width, img_height, threshold_list)

        # Calculate average time per image
        prediction_time = time.time() - start_time
        prediction_time_per_img = prediction_time/(len(os.listdir(images_dir)))

        add_parameters_to_xlsx(excel_file_path, prediction_time_per_img, model_name, step_folder, warmup_steps)



'''
************************************************************************************************************************
STEP 3: EVALUATE MODELS AND SAVE RESULTS TO EXCEL FILE
************************************************************************************************************************
'''


class BoundingBoxA:
    def __init__(self, x_min, y_min, x_max, y_max, width, height):
        self.width = width
        self.height = height
        self.x_min = x_min
        self.x_max = x_max
        self.y_min = y_min
        self.y_max = y_max


class CaptchaField:
    def __init__(self, x_min, y_min, x_max, y_max):
        self.x_min = x_min
        self.x_max = x_max
        self.y_min = y_min
        self.y_max = y_max
        self.width = x_max - x_min
        self.height = y_max - y_min


def evaluate_models_and_save_results(xml_parent_dir, actual_matrix_dir, img_dir, xlsx_file_path, threshold_list, specific_correct_list):
    def do_boxes_intersect(a, b):
        return ((abs((a.x_min + a.width / 2) - (b.x_min + b.width / 2)) * 2 < (a.width + b.width)) and
                (abs((a.y_min + a.height / 2) - (b.y_min + b.height / 2)) * 2 < (a.height + b.height)))

    def calculate_captcha_style_preds_and_accuracy(xml_dir, img_dir, specific_correct_list):
        for img_file in [file for file in os.listdir(img_dir) if file.endswith('.png')]:
            xml_file = f'{img_file[:-4]}.xml'
            # create the captcha matrix for the predictions for every image
            intersection_matrix = [[False, False, False, False],
                                   [False, False, False, False],
                                   [False, False, False, False],
                                   [False, False, False, False]]
            tree = ET.parse(os.path.join(xml_dir, xml_file))
            root = tree.getroot()
            for member in root.findall('object'):
                bndbox = member.find('bndbox')
                value = (int(bndbox.find('xmin').text),
                         int(bndbox.find('ymin').text),
                         int(bndbox.find('xmax').text),
                         int(bndbox.find('ymax').text),
                         int(bndbox.find('xmax').text) - int(bndbox.find('xmin').text),
                         int(bndbox.find('ymax').text) - int(bndbox.find('ymin').text)
                         )
                # column_name = ['xmin', 'ymin', 'xmax', 'ymax', 'width', 'height']
                a = BoundingBoxA(*value)

                # check with which captcha boxes the current bounding box intersects with and changes it in the matrix
                for ind_y, square_pos_y in enumerate([0, 30, 60, 90]):
                    for ind_x, square_pos_x in enumerate([0, 30, 60, 90]):
                        b = CaptchaField(square_pos_x, square_pos_y, square_pos_x + 30, square_pos_y + 30)
                        if intersection_matrix[ind_y][ind_x] is False:
                            if do_boxes_intersect(a, b) is True:
                                intersection_matrix[ind_y][ind_x] = True

            calculate_accuracy_for_image(intersection_matrix, actual_matrix_dir, xml_file, specific_correct_list)

    def calculate_accuracy_for_image(pred_intersection_matrix, actual_matrix_dir, xml_filename, specific_correct_list):
        txt_file = f'{xml_filename[:-4]}.txt'
        with open(f'{os.path.join(actual_matrix_dir, txt_file)}') as actual_file:
            actual_content = actual_file.read()
            actual_content = ast.literal_eval(actual_content)

            pred_accuracy = 0

            for x in range(4):
                for y in range(4):
                    if pred_intersection_matrix[x][y] == actual_content[x][y]:
                        pred_accuracy += 1

        for num in specific_correct_list:
            if isinstance(num, str):
                if num.startswith(':'):
                    specific_number = num.split(':')[1]
                    if pred_accuracy <= int(specific_number):
                        specific_correct_dict[f'number_{num}'] += 1
            elif isinstance(num, int):
                if pred_accuracy == num:
                    specific_correct_dict[f'number_{num}'] += 1

        accuracy_for_img = pred_accuracy / 16

        in_image_accuracy_list.append(accuracy_for_img)
        if accuracy_for_img == 1:
            perfect_accuracy_list.append(1)
            perfect_accuracy_name_list.append(xml_filename[:-4])
        else:
            perfect_accuracy_list.append(0)

    def add_results_to_excel_spreadsheet(xlsx_file_path, threshold, specific_correct_list, step, warmup_steps, model_name):
        in_image_accuracy = round((sum(in_image_accuracy_list) / len(in_image_accuracy_list) * 100), 2)
        perfect_accuracy = round((sum(perfect_accuracy_list) / len(perfect_accuracy_list) * 100), 2)

        wb_obj = openpyxl.load_workbook(xlsx_file_path)

        sheet = wb_obj.active

        max_row = get_max_row_with_data(sheet)
        max_column = get_max_column_with_data(sheet)

        bold_font = styles.Font(bold=True)
        standard_font = styles.Font(bold=False)

        # Check if the columns for the specific correct list values exist and creating them if needed
        for i, num in enumerate(specific_correct_list):
            add_num_to_col = True
            if isinstance(num, int):
                for column_value_index in range(1, max_column + 1):
                    if sheet.cell(row=1, column=column_value_index).value == f'sum {num}/16':
                        add_num_to_col = False
                        # if the specific correct values already have an assigned column, set the perfect_images_names two cols to the right
                        if i + 1 == len(specific_correct_list):
                            perfect_images_col = column_value_index + 2
                        break
            elif isinstance(num, str):
                formatted_num = num.split(':')[1]
                for column_value_index in range(1, max_column + 1):
                    if sheet.cell(row=1, column=column_value_index).value == f'{formatted_num} or below':
                        add_num_to_col = False
                        break

            if add_num_to_col:
                if isinstance(num, int):
                    sheet.cell(row=1, column=max_column + 1).value = f'sum {num}/16'
                    # If it's the last specific correct number, set the column for the perfect_images_names to 2 further to the right
                    if num == specific_correct_list[-1]:
                        perfect_images_col = max_column + 3
                elif isinstance(num, str):
                    sheet.cell(row=1, column=max_column + 1).value = f'{num.split(":")[1]} or below'
                sheet.cell(row=1, column=max_column + 1).font = bold_font
                wb_obj.save(os.path.join(os.getcwd(), xlsx_file_path))
                max_column = get_max_column_with_data(sheet)
                max_row = get_max_row_with_data(sheet)

        # Find the existing row for this exact model to add results
        model_name_exact = f'{model_name}_step{step}'
        for row_num in range(1, max_row + 1):
            model_name_of_sheet_line = sheet[f'A{row_num}'].value
            step_of_sheet_line = sheet[f'B{row_num}'].value
            warmup_steps_of_sheet_line = sheet[f'C{row_num}'].value
            threshold_of_sheet_line = sheet[f'D{row_num}'].value

            if model_name_of_sheet_line == model_name_exact and step_of_sheet_line == step and warmup_steps_of_sheet_line == warmup_steps and threshold_of_sheet_line == threshold:
                active_row = row_num
                break

        # Add the specific correct numbers to the Excel spreadsheet (to the correct line)
        for num in specific_correct_list:
            if isinstance(num, int):
                for column_value_index in range(1, max_column + 1):
                    if sheet.cell(row=1, column=column_value_index).value == f'sum {num}/16':
                        sheet.cell(row=active_row, column=column_value_index).value = specific_correct_dict[f'number_{num}']
                        sheet.cell(row=active_row, column=column_value_index).font = standard_font

                        break
            elif isinstance(num, str):
                formatted_num = num.split(':')[1]
                for column_value_index in range(1, max_column + 1):
                    if sheet.cell(row=1, column=column_value_index).value == f'{formatted_num} or below':
                        sheet.cell(row=active_row, column=column_value_index).value = specific_correct_dict[f'number_{num}']
                        sheet.cell(row=active_row, column=column_value_index).font = standard_font
                        break

        # Add the names of the images the model predicted perfectly to the spreadsheet
        sheet.cell(row=active_row, column=perfect_images_col).value = f'{perfect_accuracy_name_list}'
        sheet.cell(row=active_row, column=perfect_images_col).font = standard_font

        # Add accuracy and perfect accuracy to spreadsheet
        sheet.cell(row=active_row, column=6).value = in_image_accuracy
        # sheet.cell(row=active_row, column=6).font = standard_font
        sheet.cell(row=active_row, column=7).value = perfect_accuracy
        # sheet.cell(row=active_row, column=7).font = standard_font

        wb_obj.save(os.path.join(os.getcwd(), xlsx_file_path))

    for step_folder in tqdm([sf for sf in os.listdir(xml_parent_dir) if sf not in already_treated_steps]):
        for threshold in threshold_list:
            in_image_accuracy_list = []
            perfect_accuracy_list = []
            perfect_accuracy_name_list = []
            specific_correct_dict = {}
            for number in specific_correct_list:
                specific_correct_dict[f'number_{number}'] = 0

            xml_dir_with_thresh = os.path.join(xml_parent_dir, step_folder, 'predictions', f'threshold_{threshold}')
            calculate_captcha_style_preds_and_accuracy(xml_dir_with_thresh, img_dir, specific_correct_list)
            add_results_to_excel_spreadsheet(xlsx_file_path, threshold, specific_correct_list, step_folder[4:], warmup_steps, model_name)


'''
************************************************************************************************************************
STEP 4: EXECUTING DESIRED TASKS
************************************************************************************************************************
'''


def save_and_evaluate_models(saving_models, predicting_bboxes, evaluating_models):
    if saving_models:
        save_models(parent_directory, all_checkpoints, conda_environment_name, root_directory, script_name, model_name,
                    model_version, training_steps)
    if predicting_bboxes:
        print('Predicting Bounding Boxes using the different Models... ')
        predict_bounding_boxes_with_models(parent_directory, input_images_directory, threshold_list)
    if evaluating_models:
        print('Evaluating Models... ')
        evaluate_models_and_save_results(xml_parent_directory, actual_matrix_directory, input_images_directory, excel_file_path, threshold_list, specific_correct_list)


# STEP 1: SAVING THE MODELS
# already_treated_steps = ['step100', 'step1000']
already_treated_steps = []
parent_directory = os.path.join('Tensorflow', 'workspace', 'models', 'ssd_mobilenet_v2_320x320_coco17_tpu-8', 'v2', 'checkpoints_steps')
# Only to write proper checkpoint files (numbers don't matter)
all_checkpoints = ["ckpt-495", "ckpt-496", "ckpt-497", "ckpt-498", "ckpt-499", "ckpt-500"]
model_name = 'ssd_mobilenet_v2_320x320_coco17_tpu-8'
model_version = 'v2'
training_steps = 100000
warmup_steps = 100
excel_file_path = os.path.join(os.getcwd(), 'excel_files', 'NN_Optimization.xlsx')

conda_environment_name = 'venvMaturaarbeit'
root_directory = os.path.join(os.getcwd(), 'Tensorflow', 'models', 'research', 'object_detection')
script_name = 'exporter_main_v2.py'

# STEP 2: PREDICT BOUNDING BOXES
input_images_directory = os.path.join(os.getcwd(), 'train_val_test_split', 'images', 'test_bb')
threshold_list = [0, 0.1, 0.15, 0.175, 0.2, 0.225, 0.25, 0.275, 0.3, 0.4, 0.45, 0.5]


# STEP 3: EVALUATE MODELS AND SAVE RESULTS TO EXCEL FILE
xml_parent_directory = os.path.join('Tensorflow', 'workspace', 'models', 'ssd_mobilenet_v2_320x320_coco17_tpu-8', 'v2', 'checkpoints_steps')

actual_matrix_directory = os.path.join(os.getcwd(), '..', 'train_val_test_split', 'captcha_matrices')
specific_correct_list = [':4', 13, 14, 15]


# STEP 4: RUN PROGRAM (Three parameters: saving_models, predicting_bboxes and evaluating models --> can be set to True or False)
save_and_evaluate_models(True, True, True)

