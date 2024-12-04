from sklearn import svm
from sklearn.model_selection import train_test_split
import os
import cv2
import numpy as np
from PIL import Image, ImageEnhance
from tqdm import tqdm
import joblib
import xml.etree.ElementTree as ET
import matplotlib.pyplot as plt
import ast
import statistics
import openpyxl
from openpyxl import styles
import time


'''
************************************************************************************************************************
General Definitions
************************************************************************************************************************
'''


def enhance_image(image_path, contrast_enhancement_factor, brightness_enhancement_factor):
    # Open the image
    image = Image.open(image_path)

    image = image.convert('L')
    # Enhance the contrast
    enhancer = ImageEnhance.Contrast(image)
    enhanced_image = enhancer.enhance(contrast_enhancement_factor)

    enhancer = ImageEnhance.Brightness(enhanced_image)
    enhanced_image = enhancer.enhance(brightness_enhancement_factor)

    # enhanced_image = enhanced_image.convert('L')
    return enhanced_image


'''
************************************************************************************************************************
STEP 1: TRAINING SVM WITH DIFFERENT PARAMETERS
************************************************************************************************************************
'''


def train_support_vector_machine(positive_img_dir, negative_img_dir, model_saving_dir, contrast_factor, brightness_factor, hog_par_list):
    def create_positive_and_negative_sample_lists(positive_img_dir, negative_img_dir):
        positive = []
        negative = []
        labels_pos_neg = []

        # add positive images to the positive list
        for positive_img in os.listdir(positive_img_dir):
            pos_img_path = os.path.join(positive_img_dir, positive_img)
            labels_pos_neg.append(pos_img_path)
            img_enhanced_pos = np.array(enhance_image(pos_img_path, contrast_factor, brightness_factor))
            # cv2.imshow('window', img_enhanced_pos)
            cv2.waitKey(0)
            cv2.destroyAllWindows()
            # pos_img = cv2.imread(img_enhanced)
            # Check the image type
            if img_enhanced_pos.dtype != np.uint8 or (img_enhanced_pos.ndim != 2 and img_enhanced_pos.ndim != 3):
                # Convert image to appropriate type
                img_enhanced_pos = cv2.cvtColor(img_enhanced_pos,
                                                cv2.COLOR_BGR2GRAY) if img_enhanced_pos.ndim == 3 else img_enhanced_pos.astype(
                    np.uint8)
            positive.append(img_enhanced_pos)

        # add negative images to the negative list
        for negative_img in os.listdir(negative_img_dir):
            neg_img_path = os.path.join(negative_img_dir, negative_img)
            labels_pos_neg.append(neg_img_path)
            img_enhanced_neg = np.array(enhance_image(neg_img_path, contrast_factor, brightness_factor))
            # cv2.imshow('window', img_enhanced_neg)
            cv2.waitKey(0)
            cv2.destroyAllWindows()
            # neg_img = cv2.imread(neg_img_path)
            # Check the image type
            if img_enhanced_neg.dtype != np.uint8 or (img_enhanced_neg.ndim != 2 and img_enhanced_neg.ndim != 3):
                # Convert image to appropriate type
                img_enhanced_neg = cv2.cvtColor(img_enhanced_neg,
                                                cv2.COLOR_BGR2GRAY) if img_enhanced_neg.ndim == 3 else img_enhanced_neg.astype(
                    np.uint8)
            negative.append(img_enhanced_neg)

        return positive, negative, labels_pos_neg

    def compute_hog_features_for_images(images, labels_pos_neg, hog_param_list):
        data = []
        # Define HOG parameters
        win_size = (hog_param_list[0], hog_param_list[0])
        block_size = (hog_param_list[1], hog_param_list[1])
        block_stride = (hog_param_list[2], hog_param_list[2])
        cell_size = (hog_param_list[3], hog_param_list[3])
        nbins = hog_param_list[4]

        counter = 0

        hog = cv2.HOGDescriptor(win_size, block_size, block_stride, cell_size, nbins)
        for img in images:
            # Check if the image has a third dimension (color channels)
            if len(img.shape) == 3 and img.shape[2] == 4:
                img = img[:, :, :3]  # Keep only the first three channels (RGB/BGR)
            elif len(img.shape) == 2:
                img = cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)  # Convert grayscale to BGR

            if img.shape[:1] != win_size[0]:
                img.resize(win_size[0], win_size[1])

            # print(labels_pos_neg[counter])
            counter += 1
            features = hog.compute(img)
            data.append(features.flatten())

        return data

    def train_svm(train_x, train_y, model_saving_dir):
        print('training support vector machine...')
        clf = svm.SVC(kernel='rbf')
        clf.fit(train_x, train_y)

        # save the model for usage without retraining
        svm_model_name = f'svm_model_{hog_par_list[0]}_{hog_par_list[1]}_{hog_par_list[2]}_{hog_par_list[3]}_{hog_par_list[4]}_c{contrast_factor}_b{brightness_factor}.pkl'
        # os.makedirs(os.path.join(os.path.join(model_saving_dir, svm_model_name)), exist_ok=True)
        joblib.dump(clf, f'{os.path.join(model_saving_dir, svm_model_name)}')
        return clf

    # Run preparation steps
    positive, negative, labels_posit_negat = create_positive_and_negative_sample_lists(positive_img_dir, negative_img_dir)
    images = positive + negative
    labels = ([1] * len(positive)) + ([0] * len(negative))

    # Compute Hog features
    data = compute_hog_features_for_images(images, labels_posit_negat, hog_par_list)
    # Convert data and labels to numpy arrays
    data = np.array(data, dtype=np.float32)
    labels = np.array(labels, dtype=np.int32)

    # Split train_val_test_split into train_bb and test_bb portions
    train_x, test_x, train_y, test_y = train_test_split(data, labels, test_size=0.2, random_state=42)

    start_time = time.time()

    # Train and save Svm model
    svm_model = train_svm(train_x, train_y, model_saving_dir)

    elapsed_time = time.time() - start_time
    print(f'Training time: {elapsed_time} seconds')


'''
************************************************************************************************************************
PREDICT CROSSWALKS USING THE TRAINED SVM MODEL
************************************************************************************************************************
'''


def predict_crosswalks_using_svm(image_dir, output_dir_xmls, hog_par_list, contrast_factor, brightness_factor, sliding_window_stride, threshold_list, scale_range):
    def compute_hog_features(image):
        # Define HOG parameters
        win_size = (hog_par_list[0], hog_par_list[0])
        block_size = (hog_par_list[1], hog_par_list[1])
        block_stride = (hog_par_list[2], hog_par_list[2])
        cell_size = (hog_par_list[3], hog_par_list[3])
        nbins = hog_par_list[4]

        hog = cv2.HOGDescriptor(win_size, block_size, block_stride, cell_size, nbins)
        hog_features = hog.compute(image)
        hog_features = hog_features.flatten()
        return hog_features

    def sliding_window(image, clf, win_size, stride, threshold_list):
        detections = {}
        for thresh in threshold_list:
            detections[f'detections_{thresh}'] = []
        height, width = image.shape[:2]

        for y in range(0, height - win_size[1], stride):
            for x in range(0, width - win_size[0], stride):
                window = image[y:y + win_size[1], x:x + win_size[0]]
                hog_features = compute_hog_features(window)  # Compute HOG features for the window
                decision = clf.decision_function(hog_features.reshape(1, -1))[0]  # Decision function value
                # if decision > 0:
                #     print('+', decision)
                # else:
                #     print(decision)
                for threshold in threshold_list:
                    if decision > threshold:
                        detections[f'detections_{threshold}'].append((x, y, x + win_size[0], y + win_size[1]))  # Store the bounding box coordinates

        return detections

    def multi_scale_sliding_window(image, clf, win_size, stride, threshold_list, scale_range):
        detections = {}
        for thresh in threshold_list:
            detections[f'detections_{thresh}'] = []
        height, width = image.shape[:2]

        for scale in scale_range:
            # print(scale)
            resized_image = cv2.resize(image, (int(width * scale), int(height * scale)))
            scaled_detections = sliding_window(resized_image, clf, win_size, stride, threshold_list)
            # Scale detections back to the original image size
            for scaled_detections_thresh_key in scaled_detections:
                threshold = scaled_detections_thresh_key.split('_')[1]
                scaled_detections_thresh_list = scaled_detections[f'detections_{threshold}']
                for (x1, y1, x2, y2) in scaled_detections_thresh_list:
                    x1 = int(x1 / scale)
                    y1 = int(y1 / scale)
                    x2 = int(x2 / scale)
                    y2 = int(y2 / scale)
                    detections[f'detections_{threshold}'].append((x1, y1, x2, y2))
        return detections

    def save_bounding_boxes_to_xml(image_name, detections, output_dir):
        for detections_key in detections:
            detections_list = detections[detections_key]
            threshold = detections_key.split('_')[1]

            # Create root element
            root = ET.Element("annotation")
            filename = ET.SubElement(root, "filename")
            filename.text = image_name

            # Create object elements for each bounding box
            for idx, (x1, y1, x2, y2) in enumerate(detections_list, start=1):
                obj = ET.SubElement(root, "object")
                obj_name = ET.SubElement(obj, "name")
                obj_name.text = f"object_{idx}"
                bbox = ET.SubElement(obj, "bndbox")
                xmin = ET.SubElement(bbox, "xmin")
                xmin.text = str(x1)
                ymin = ET.SubElement(bbox, "ymin")
                ymin.text = str(y1)
                xmax = ET.SubElement(bbox, "xmax")
                xmax.text = str(x2)
                ymax = ET.SubElement(bbox, "ymax")
                ymax.text = str(y2)

            # Create XML tree and write to file
            tree = ET.ElementTree(root)
            xml_filename = os.path.splitext(image_name)[0] + ".xml"
            xml_path = os.path.join(output_dir, f'threshold_{threshold}', xml_filename)
            os.makedirs(os.path.join(output_dir, f'threshold_{threshold}'), exist_ok=True)
            tree.write(xml_path)

    print('Predicting...')
    svm_model_name = os.path.join(model_saving_directory, f'svm_model_{hog_par_list[0]}_{hog_par_list[1]}_{hog_par_list[2]}_{hog_par_list[3]}_{hog_par_list[4]}_c{contrast_factor}_b{brightness_factor}.pkl')
    clf = joblib.load(svm_model_name)

    win_size = hog_par_list[0]

    # image_dir = os.path.join(os.getcwd(), '..', 'train_val_test_split', 'Crosswalk')
    # output_dir_xmls = os.path.join(os.getcwd(), 'Optimizing_Labels', f'{svm_model_name}')

    # Record the start time
    start_time = time.time()

    for image in tqdm([file for file in os.listdir(image_dir) if file.endswith('.png')]):
        image_path = os.path.join(image_dir, image)
        # image_open = cv2.imread(image_path)
        enhanced_image = np.array(enhance_image(image_path, contrast_factor, brightness_factor))

        detections = multi_scale_sliding_window(np.array(enhanced_image), clf, (win_size, win_size), sliding_window_stride, threshold_list, scale_range)

        # save detections to .xml file
        save_bounding_boxes_to_xml(image, detections, output_dir_xmls)

    elapsed_time = time.time() - start_time
    print(f'total time for {svm_model_name}: {elapsed_time} seconds')
    time_per_image = elapsed_time / len(os.listdir(image_dir))
    print('len test_dir', len(os.listdir(image_dir)))
    print(f'time per image for {svm_model_name}: {time_per_image} seconds')


'''
************************************************************************************************************************
EVALUATION OF MODELS AND SAVING RESULTS
************************************************************************************************************************
'''

class BoundingBoxA:
    def __init__(self, x_min, y_min, x_max, y_max, width, height):
        self.width = width
        self.height = height
        self.x_min = x_min
        self.x_max = x_max
        self.y_min = y_min
        self.y_max = y_max


class CaptchaField:
    def __init__(self, x_min, y_min, x_max, y_max):
        self.x_min = x_min
        self.x_max = x_max
        self.y_min = y_min
        self.y_max = y_max
        self.width = x_max - x_min
        self.height = y_max - y_min


def evaluate_models_and_save_results(xml_dir, actual_matrix_dir, img_dir, excel_filename, hog_par_list, contrast_factor, brightness_factor, threshold_list, specific_correct_list):
    def do_boxes_intersect(a, b):
        return ((abs((a.x_min + a.width / 2) - (b.x_min + b.width / 2)) * 2 < (a.width + b.width)) and
                (abs((a.y_min + a.height / 2) - (b.y_min + b.height / 2)) * 2 < (a.height + b.height)))

    def calculate_captcha_style_preds_and_accuracy(xml_dir, img_dir, specific_correct_list):
        for img_file in tqdm([file for file in os.listdir(img_dir) if file.endswith('.png')]):
            xml_file = f'{img_file[:-4]}.xml'
            intersection_matrix = [[False, False, False, False],
                                   [False, False, False, False],
                                   [False, False, False, False],
                                   [False, False, False, False]]
            tree = ET.parse(os.path.join(xml_dir, xml_file))
            root = tree.getroot()
            for member in root.findall('object'):
                bndbox = member.find('bndbox')
                value = (int(bndbox.find('xmin').text),
                         int(bndbox.find('ymin').text),
                         int(bndbox.find('xmax').text),
                         int(bndbox.find('ymax').text),
                         int(bndbox.find('xmax').text) - int(bndbox.find('xmin').text),
                         int(bndbox.find('ymax').text) - int(bndbox.find('ymin').text)
                         )
                # column_name = ['xmin', 'ymin', 'xmax', 'ymax', 'width', 'height']
                a = BoundingBoxA(*value)

                # check with which captcha boxes the current bounding box intersects with and changes it in the matrix
                for ind_y, square_pos_y in enumerate([0, 30, 60, 90]):
                    for ind_x, square_pos_x in enumerate([0, 30, 60, 90]):
                        b = CaptchaField(square_pos_x, square_pos_y, square_pos_x + 30, square_pos_y + 30)
                        if intersection_matrix[ind_y][ind_x] is False:
                            if do_boxes_intersect(a, b) is True:
                                intersection_matrix[ind_y][ind_x] = True

            calculate_accuracy_for_image(intersection_matrix, actual_matrix_dir, xml_file, specific_correct_list)

    def calculate_accuracy_for_image(pred_intersection_matrix, actual_matrix_dir, xml_filename, specific_correct_list):
        txt_file = f'{xml_filename[:-4]}.txt'
        with open(f'{os.path.join(actual_matrix_dir, txt_file)}') as actual_file:
            actual_content = actual_file.read()
            actual_content = ast.literal_eval(actual_content)

            pred_accuracy = 0

            for x in range(4):
                for y in range(4):
                    if pred_intersection_matrix[x][y] == actual_content[x][y]:
                        pred_accuracy += 1

        for num in specific_correct_list:
            if isinstance(num, str):
                if num.startswith(':'):
                    specific_number = num.split(':')[1]
                    if pred_accuracy <= int(specific_number):
                        specific_correct_dict[f'number_{num}'] += 1
            elif isinstance(num, int):
                if pred_accuracy == num:
                    specific_correct_dict[f'number_{num}'] += 1

        accuracy_for_img = pred_accuracy / 16

        in_image_accuracy_list.append(accuracy_for_img)
        if accuracy_for_img == 1:
            perfect_accuracy_list.append(1)
            perfect_accuracy_name_list.append(xml_filename[:-4])
        else:
            perfect_accuracy_list.append(0)

    def add_results_to_excel_spreadsheet(xlsx_filename, hog_par_list, contrast_factor, brightness_factor, threshold, specific_correct_list):
        in_image_accuracy = round((sum(in_image_accuracy_list) / len(in_image_accuracy_list) * 100), 2)
        perfect_accuracy = round((sum(perfect_accuracy_list) / len(perfect_accuracy_list) * 100), 2)

        wb_obj = openpyxl.load_workbook(xlsx_filename)

        sheet = wb_obj.active

        row = sheet.max_row
        column = sheet.max_column

        bold_font = styles.Font(bold=True)

        # Check if the columns for the specific correct list values exist and creating them if needed
        for num in specific_correct_list:
            add_num_to_col = True
            if isinstance(num, int):
                for column_value_index in range(1, column+1):
                    if sheet.cell(row=1, column=column_value_index).value == f'sum {num}/16':
                        add_num_to_col = False
                        break
            elif isinstance(num, str):
                formatted_num = num.split(':')[1]
                for column_value_index in range(1, column+1):
                    if sheet.cell(row=1, column=column_value_index).value == f'{formatted_num} or below':
                        add_num_to_col = False
                        break

            if add_num_to_col:
                if isinstance(num, int):
                    sheet.cell(row=1, column=column + 1).value = f'sum {num}/16'
                elif isinstance(num, str):
                    sheet.cell(row=1, column=column + 1).value = f'{num.split(":")[1]} or below'
                sheet.cell(row=1, column=column + 1).font = bold_font
                wb_obj.save(os.path.join(os.getcwd(), xlsx_filename))
                column = sheet.max_column
                row = sheet.max_row

        for num in specific_correct_list:
            if isinstance(num, int):
                for column_value_index in range(1, column+1):
                    if sheet.cell(row=1, column=column_value_index).value == f'sum {num}/16':
                        sheet.cell(row=row+1, column=column_value_index).value = specific_correct_dict[f'number_{num}']
                        break
            elif isinstance(num, str):
                formatted_num = num.split(':')[1]
                for column_value_index in range(1, column+1):
                    if sheet.cell(row=1, column=column_value_index).value == f'{formatted_num} or below':
                        sheet.cell(row=row+1, column=column_value_index).value = specific_correct_dict[f'number_{num}']
                        break

        sheet[f'A{row + 1}'].value = f'svm_model_{hog_par_list[0]}_{hog_par_list[1]}_{hog_par_list[2]}_{hog_par_list[3]}_{hog_par_list[4]}_c{contrast_factor}_b{brightness_factor}.pkl'
        sheet[f'B{row + 1}'].value = hog_par_list[0]
        sheet[f'C{row + 1}'].value = hog_par_list[1]
        sheet[f'D{row + 1}'].value = hog_par_list[2]
        sheet[f'E{row + 1}'].value = hog_par_list[3]
        sheet[f'F{row + 1}'].value = hog_par_list[4]
        sheet[f'G{row + 1}'].value = contrast_factor
        sheet[f'H{row + 1}'].value = brightness_factor
        sheet[f'I{row + 1}'].value = threshold
        sheet[f'J{row + 1}'].value = in_image_accuracy
        sheet[f'K{row + 1}'].value = perfect_accuracy

        sheet.cell(row=row+1, column=11+len(specific_correct_list)+2).value = f'{perfect_accuracy_name_list}'

        wb_obj.save(os.path.join(os.getcwd(), xlsx_filename))

    print('Evaluating Model...')

    for threshold in threshold_list:
        in_image_accuracy_list = []
        perfect_accuracy_list = []
        perfect_accuracy_name_list = []
        specific_correct_dict = {}
        for number in specific_correct_list:
            specific_correct_dict[f'number_{number}'] = 0

        xml_dir_with_thresh = os.path.join(xml_dir, f'threshold_{threshold}')
        calculate_captcha_style_preds_and_accuracy(xml_dir_with_thresh, img_dir, specific_correct_list)
        add_results_to_excel_spreadsheet(excel_filename, hog_par_list, contrast_factor, brightness_factor, threshold, specific_correct_list)


'''
************************************************************************************************************************
EXECUTING THE DESIRED TASKS
************************************************************************************************************************
'''


def find_best_hog_model_parameters(list_step1, list_step2, list_step3, train_svm, predict_bboxes, evaluate_models):
    # STEP 1: Training SVM
    # Extracting variables from input list
    contrast_factor_list = list_step1[0]
    brightness_factor_list = list_step1[1]
    win_size_list = list_step1[2]
    block_size_list = list_step1[3]
    block_stride_list = list_step1[4]
    cell_size_list = list_step1[5]
    nbins = list_step1[6]
    model_saving_dir = list_step1[7]

    # STEP 2: Predicting using SVM
    test_image_dir = list_step2[0]
    output_dir_xmls = list_step2[1]
    sliding_window_stride = list_step2[2]
    threshold_list = list_step2[3]
    scale_range = list_step2[4]
    specific_correct_list = list_step2[5]

    # STEP 3: Evaluating the Models
    actual_matrix_dir = list_step3[0]
    xlsx_file_path = list_step3[1]

    amount_of_models = len(win_size_list) * len(block_size_list) * len(block_stride_list) * len(cell_size_list) * len(contrast_factor_list) * len(brightness_factor_list)
    amount_of_models_trained = 0

    for contr_fact in contrast_factor_list:
        for bright_fact in brightness_factor_list:
            for win_size in win_size_list:
                for block_size in block_size_list:
                    for block_stride in block_stride_list:
                        for cell_size in cell_size_list:
                            amount_of_models_trained += 1
                            print(f'''
(Model {amount_of_models_trained}/{amount_of_models})''')
                            hog_par_list = [win_size, block_size, block_stride, cell_size, nbins]
                            if train_svm:
                                print(hog_par_list)
                                train_support_vector_machine(positive_image_directory, negative_image_directory, model_saving_dir, contr_fact, bright_fact, hog_par_list)

                            output_dir_xmls_excact = os.path.join(output_dir_xmls,
                                                                  f'svm_model_{hog_par_list[0]}_{hog_par_list[1]}_{hog_par_list[2]}_{hog_par_list[3]}_{hog_par_list[4]}_c{contr_fact}_b{bright_fact}')
                            if predict_bboxes:
                                predict_crosswalks_using_svm(test_image_dir, output_dir_xmls_excact, hog_par_list, contr_fact, bright_fact, sliding_window_stride, threshold_list, scale_range)
                            if evaluate_models:
                                evaluate_models_and_save_results(output_dir_xmls_excact, actual_matrix_dir, test_image_dir, xlsx_file_path, hog_par_list, contr_fact, bright_fact, threshold_list, specific_correct_list)


# STEP 1:
positive_image_directory = os.path.join(os.getcwd(), 'train_val_test_split', 'images', 'square_crosswalk_regions')
negative_image_directory = os.path.join(os.getcwd(), 'train_val_test_split', 'images', 'square_non_crosswalk_regions')

# contrast_factor_list = [1, 2, 3, 5, 7, 10]
contrast_factor_list = [3]
# brightness_factor_list = [0.1, 0.05, 0.01, 0.005]
brightness_factor_list = [0.1]

win_size_list = [32]
# block_size_list = [8, 16]
block_size_list = [8]
# block_stride_list = [8, 4]
block_stride_list = [8]
# cell_size_list = [8, 4]
cell_size_list = [4]
nbins = 9

model_saving_directory = os.path.join(os.getcwd(), 'Optimizing', 'trained_models_poly')

list_step1 = [contrast_factor_list, brightness_factor_list, win_size_list, block_size_list, block_stride_list, cell_size_list, nbins, model_saving_directory]

# STEP 2:
sliding_window_stride = 8
# threshold_list = [-1, -0.5, -0.25, 0, 0.25, 0.5, 0.75, 1]
threshold_list = [1]
scale_range = [2, 1.5]
test_image_directory = os.path.join(os.getcwd(), '..', 'dataset', 'train_val_test_split', 'images', 'test')
output_directory_xmls = os.path.join(os.getcwd(), 'Optimizing', 'Predicted_Labels_rbf')
specific_correct_list = [':4', 13, 14, 15]

list_step2 = [test_image_directory, output_directory_xmls, sliding_window_stride, threshold_list, scale_range, specific_correct_list]

actual_matrix_directory = os.path.join(os.getcwd(), '..', 'dataset', 'captcha_matrices')
excel_file_path = os.path.join(os.getcwd(), 'Optimizing', 'HOG_Optimization_Layout_copy.xlsx')

# STEP 3:
list_step3 = [actual_matrix_directory, excel_file_path]

# RUNNING THE SCRIPT
find_best_hog_model_parameters(list_step1, list_step2, list_step3, True, True, True)

