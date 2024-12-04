import ast
import os
import subprocess
import pprint
import time

import matplotlib
import ast
import numpy as np
import matplotlib.pyplot as plt
import cv2
from PIL import Image

from tqdm import tqdm
import openpyxl
from openpyxl import styles

import absl.logging

from IPython import display

import tensorflow as tf
import tensorflow_datasets as tfds


import orbit
import tensorflow_models as tfm
from official.vision.data import tfrecord_lib
# from official.vision.utils import summary_manager
from official.vision.serving import export_saved_model_lib
from official.vision.utils.object_detection import visualization_utils

pp = pprint.PrettyPrinter(indent=4)  # Set Pretty Print Indentation
# print(tf.__version__) # Check the version of tensorflow used


'''
************************************************************************************************************************
STEP: GENERAL DEFINITIONS
************************************************************************************************************************
'''


def get_max_column_with_data(sheet):
    max_column = 0
    # Loop through all rows to check each row's actual max column with data
    for row in sheet.iter_rows():
        for cell in reversed(row):  # Start checking from the last column
            if cell.value is not None:
                max_column = max(max_column, cell.column)
                break  # Stop at the first non-empty cell in the row
    return max_column


def get_max_row_with_data(sheet):
    max_row = 0
    # Loop through all rows to check each row's actual max column with data
    for column in sheet.iter_cols():
        for cell in reversed(column):  # Start checking from the last column
            if cell.value is not None:
                max_row = max(max_row, cell.row)
                break  # Stop at the first non-empty cell in the row
    return max_row


def config_model():
    exp_config = tfm.core.exp_factory.get_exp_config('mnv2_deeplabv3_pascal')

    if not os.path.exists(model_ckpt_path):
        os.mkdir(model_ckpt_path)

        subprocess.run(["gsutil", "cp", "gs://tf_model_garden/cloud/vision-2.0/deeplab/deeplabv3_mobilenetv2_coco"
                                        "/best_ckpt-0.data-00000-of-00001/", model_ckpt_path])
        subprocess.run(
          ["gsutil", "cp", "gs://tf_model_garden/cloud/vision-2.0/deeplab/deeplabv3_mobilenetv2_coco/best_ckpt-0.index/",
           model_ckpt_path])

    # Backbone Config
    exp_config.task.init_checkpoint = model_ckpt_path + 'ckpt-0'
    exp_config.task.freeze_backbone = True

    # Model Config
    exp_config.task.model.num_classes = num_classes
    exp_config.task.model.input_size = input_size

    # Training data Config
    exp_config.task.train_data.aug_scale_min = 1.0
    exp_config.task.train_data.aug_scale_max = 1.0
    exp_config.task.train_data.input_path = train_data_tfrecords
    exp_config.task.train_data.global_batch_size = BATCH_SIZE
    exp_config.task.train_data.dtype = 'float32'
    exp_config.task.train_data.output_size = [HEIGHT, WIDTH]
    exp_config.task.train_data.preserve_aspect_ratio = False
    exp_config.task.train_data.seed = 21 # Reproducable Training data

    # Validation data Config
    exp_config.task.validation_data.input_path = val_data_tfrecords
    exp_config.task.validation_data.global_batch_size = BATCH_SIZE
    exp_config.task.validation_data.dtype = 'float32'
    exp_config.task.validation_data.output_size = [HEIGHT, WIDTH]
    exp_config.task.validation_data.preserve_aspect_ratio = False
    exp_config.task.validation_data.groundtruth_padded_size = [HEIGHT, WIDTH]
    exp_config.task.validation_data.seed = 21 # Reproducable Validation data
    exp_config.task.validation_data.resize_eval_groundtruth = True # To enable validation loss

    return exp_config


def config_trainer():
    logical_device_names = [logical_device.name
                            for logical_device in tf.config.list_logical_devices()]

    if 'GPU' in ''.join(logical_device_names):
        print('Device set to GPU')
        device = 'GPU'
    elif 'TPU' in ''.join(logical_device_names):
        print('This may be broken in Colab.')
        device = 'TPU'
    else:
        print('Running on CPU is slow, so only train_bb for a few steps.')
        device = 'CPU'

    exp_config.trainer.steps_per_loop = 100  # int(amount_of_train_samples // BATCH_SIZE)

    exp_config.trainer.summary_interval = exp_config.trainer.steps_per_loop  # steps_per_loop = num_of_validation_examples // eval_batch_size
    exp_config.trainer.checkpoint_interval = 100  # exp_config.trainer.steps_per_loop
    exp_config.trainer.validation_interval = exp_config.trainer.steps_per_loop
    exp_config.trainer.validation_steps = 100  # int(amount_of_train_samples // BATCH_SIZE)  # validation_steps = num_of_validation_examples // eval_batch_size
    exp_config.trainer.train_steps = train_steps
    exp_config.trainer.optimizer_config.warmup.linear.warmup_steps = warmup_steps
    exp_config.trainer.optimizer_config.learning_rate.type = 'cosine'
    exp_config.trainer.optimizer_config.learning_rate.cosine.decay_steps = train_steps
    exp_config.trainer.optimizer_config.learning_rate.cosine.initial_learning_rate = 0.1
    exp_config.trainer.optimizer_config.warmup.linear.warmup_learning_rate = 0.05

    # KEEP FOR DEBUGGING!!!
    # pp.pprint(exp_config.as_dict())

    # Setting up the Strategy
    if exp_config.runtime.mixed_precision_dtype == tf.float16:
        tf.keras.mixed_precision.set_global_policy('mixed_float16')

    if 'GPU' in ''.join(logical_device_names):
        distribution_strategy = tf.distribute.MirroredStrategy()
    elif 'TPU' in ''.join(logical_device_names):
        tf.tpu.experimental.initialize_tpu_system()
        tpu = tf.distribute.cluster_resolver.TPUClusterResolver(tpu='/device:TPU_SYSTEM:0')
        distribution_strategy = tf.distribute.experimental.TPUStrategy(tpu)
    else:
        print('Warning: this will be really slow.')
        distribution_strategy = tf.distribute.OneDeviceStrategy(logical_device_names[0])

    print("Trainer is configured.")

    '''
    Create the Task object (tfm.core.base_task.Task) from the config_definitions.TaskConfig.
    The Task object has all the methods necessary for building the dataset, building the model, and running training & evaluation.
    These methods are driven by tfm.core.train_lib.run_experiment.
    '''

    with distribution_strategy.scope():
        task = tfm.core.task_factory.get_task(exp_config.task, logging_dir=model_dir)

    # Visualize a batch of the data.
    # Check if the dataset is not empty
    # Visualize a few examples

    # KEEP FOR DEBUGGING!!!
    # for images, masks in task.build_inputs(exp_config.task.train_data).take(1):
    #     print()
    #     print(f'images.shape: {str(images.shape):16}  images.dtype: {images.dtype!r}')
    #     print(f'masks.shape: {str(masks["masks"].shape):16} images.dtype: {masks["masks"].dtype!r}')

    return distribution_strategy, task


# def plot_masks(display_list):
#     plt.figure(figsize=(10, 6))
#
#     title = ['Input Image', 'True Mask', 'Predicted Mask']
#
#     for i in range(len(display_list)):
#         plt.subplot(1, len(display_list), i+1)
#         plt.title(title[i])
#         plt.imshow(tf.keras.utils.array_to_img(display_list[i]))
#
#         plt.axis('off')
#     plt.show()
#
#
# print(matplotlib.get_backend())
#
# num_examples = 0
#
# for images, masks in task.build_inputs(exp_config.task.train_data).take(num_examples):
#     plot_masks([images[0], masks['masks'][0]])


def train_model():
    model, eval_logs = tfm.core.train_lib.run_experiment(
        distribution_strategy=distribution_strategy,
        task=task,
        mode='train_and_eval',
        params=exp_config,
        model_dir=model_dir,
        # eval_summary_manager=summary_manager.maybe_build_eval_summary_manager(
        #           params=exp_config, model_dir=model_dir),
        run_post_eval=True)


'''
************************************************************************************************************************
SAVING THE MODELS FROM CHECKPOINT FILES (--> saved_model.pb file)
************************************************************************************************************************
'''


def save_models():
    def create_checkpoint_file(parent_dir, step_n_dir, all_checkpoints):
        checkpoint_content = f'model_checkpoint_path: "ckpt-{int(int(step_n_dir[4:]))}"\n'

        for checkpoint in all_checkpoints:
            checkpoint_content += f'all_model_checkpoint_paths: "{checkpoint}"\n'

        # Write the content to a file named "checkpoint"
        checkpoint_file_path = os.path.join(parent_dir, step_n_dir, 'checkpoint')
        os.makedirs(os.path.dirname(checkpoint_file_path), exist_ok=True)

        checkpoint_file = open(checkpoint_file_path, 'w')

        checkpoint_file.write(checkpoint_content)

        checkpoint_file.close()

        print(f'Checkpoint file created at: {checkpoint_file_path}')

    def export_model(ckpt_step_exact, exported_model_dir):
        export_saved_model_lib.export_inference_graph(
            input_type='image_tensor',
            batch_size=1,
            input_image_size=[HEIGHT, WIDTH],
            params=exp_config,
            checkpoint_path=tf.train.latest_checkpoint(ckpt_step_exact),
            export_dir=exported_model_dir)

    for step_dir in os.listdir(ckpt_parent_dir):
        # Create a checkpoint file
        create_checkpoint_file(ckpt_parent_dir, step_dir, all_checkpoints)
        # Make "saved_model" directory in the folder if it doesn't exist already
        os.makedirs(os.path.join(ckpt_parent_dir, step_dir, 'saved_model'), exist_ok=True)
    for step_dir in os.listdir(ckpt_parent_dir):
        print(os.path.join(os.getcwd(), ckpt_parent_dir, step_dir, 'saved_model', 'saved_model.pb'))
        if not os.path.exists(os.path.join(os.getcwd(), ckpt_parent_dir, step_dir, 'saved_model', 'saved_model.pb')):
            export_model(os.path.join(ckpt_parent_dir, step_dir), os.path.join(ckpt_parent_dir, step_dir, 'saved_model'))


def get_predicted_masks_from_model():
    def import_model(exported_model_dir):
        imported = tf.saved_model.load(exported_model_dir)
        model_fn = imported.signatures['serving_default']
        return model_fn

    # Function to display original image, original mask, and predicted mask
    def plot_masks(images, titles):
        fig, axs = plt.subplots(1, len(images), figsize=(15, 5))
        for i, img in enumerate(images):
            axs[i].imshow(img)
            axs[i].axis('off')
            axs[i].set_title(titles[i])  # Add title to each image
        plt.show()

    def create_mask(predicted_mask_logits):
        predicted_mask = tf.argmax(predicted_mask_logits, axis=-1)
        predicted_mask = tf.squeeze(predicted_mask, axis=0)  # Remove batch dimension
        return predicted_mask.numpy()

    def save_pred_binary_mask(predicted_mask, parent_mask_out_dir, step_folder):
        os.makedirs(os.path.join(parent_mask_out_dir, step_folder, 'predictions'), exist_ok=True)
        pred_mask_out_path = os.path.join(parent_mask_out_dir, step_folder, 'predictions', image)

        cv2.imwrite(pred_mask_out_path, predicted_mask)

    def add_parameters_to_xlsx(avg_time_per_img, step_folder):
        wb_obj = openpyxl.load_workbook(xlsx_file_path)

        sheet = wb_obj.active

        row = get_max_row_with_data(sheet)

        sheet[f'A{row + 1}'].value = f'{model_name}_{step_folder}'
        sheet[f'B{row + 1}'].value = int(step_folder[4:])
        sheet[f'C{row + 1}'].value = warmup_steps
        sheet[f'D{row + 1}'].value = avg_time_per_img

        wb_obj.save(xlsx_file_path)

    def predict_for_image(parent_img_dir, parent_mask_dir, filename, model_fn):
        # Load the original image and mask
        original_image_path = os.path.join(parent_img_dir, filename)
        original_mask_path = os.path.join(parent_mask_dir, f'{filename[:-4]}_trimap.png')

        # Load and preprocess image
        original_image = cv2.imread(original_image_path)  # Or however you load the image
        original_image = cv2.cvtColor(original_image, cv2.COLOR_BGR2RGB)  # Convert BGR to RGB
        resized_image = cv2.resize(original_image, (WIDTH, HEIGHT))  # Resize to model's input size

        # Load and preprocess mask
        original_mask = cv2.imread(original_mask_path, cv2.IMREAD_GRAYSCALE)  # Load grayscale mask
        resized_mask = cv2.resize(original_mask, (WIDTH, HEIGHT))  # Resize to the same size as the image

        # Prepare image for the model
        input_image = tf.cast(resized_image, dtype=tf.uint8)
        input_image = tf.expand_dims(input_image, axis=0)  # Add batch dimension

        # Predict the mask using the trained model (replace `model_fn` with your actual model)
        predicted_mask_logits = model_fn(input_image)  # Assuming the output is logits

        # Convert logits to a final mask
        predicted_mask = create_mask(predicted_mask_logits['logits'])

        if plot_predictions:
            # Titles for the images
            titles = ['Input Image', 'True Mask', 'Predicted Mask']

            # Display the original image, original mask, and the predicted mask with titles
            plot_masks([resized_image, resized_mask, predicted_mask], titles)

        if save_pred_binary_masks:
            save_pred_binary_mask(predicted_mask, ckpt_parent_dir, step_dir)

    # Sorting them by amount of training steps rather than alphabetically
    steps_list = []
    for step_folder in os.listdir(ckpt_parent_dir):
        steps_list.append(int(step_folder[4:]))
    steps_list.sort()
    for i, step in enumerate(steps_list):
        steps_list[i] = f'step{step}'

    for step_dir in tqdm(steps_list):
        if not os.path.exists(os.path.join(os.getcwd(), ckpt_parent_dir, step_dir, 'predictions')):
            # load the model
            model_fn = import_model(os.path.join(ckpt_parent_dir, step_dir, 'saved_model'))
            # Initialize a Start timer to get average time for image
            start_time = time.time()
            # Process images in the test_bb directory
            for i, image in enumerate(os.listdir(test_img_dir)):
                if i > -1:  # To limit the amount of results you see/save
                    predict_for_image(test_img_dir, test_mask_dir, image, model_fn)
            prediction_time = time.time() - start_time
            prediction_time_per_img = prediction_time/len(os.listdir(test_img_dir))
            add_parameters_to_xlsx(prediction_time_per_img, step_dir)


def evaluate_models():
    def load_binary_mask(mask_path):
        """
        Load the binary mask from a PNG image using PIL.

        Args:
            mask_path: Path to the PNG image of the mask.

        Returns:
            2D numpy array where the pixel values correspond directly to the image file.
        """
        # Open the mask image with PIL
        mask = Image.open(mask_path)

        # Convert the image to a numpy array
        binary_mask = np.array(mask)

        return binary_mask

    def check_crosswalk_in_grid(predicted_mask, grid_size=(4, 4)):
        """
        Check if crosswalk pixels exist in each grid cell of the image.

        Args:
            predicted_mask: 2D numpy array of the predicted mask (binary: 1 for crosswalk, 0 for background).
            grid_size: Tuple indicating the grid size (default is 4x4).

        Returns:
            4x4 boolean matrix where True indicates the presence of crosswalk pixels in that grid square.
        """
        # Get dimensions of the mask
        height, width = predicted_mask.shape

        # Calculate the size of each grid square
        grid_height = height // grid_size[0]
        grid_width = width // grid_size[1]

        # Initialize a 4x4 boolean matrix
        crosswalk_grid = np.zeros(grid_size, dtype=bool)

        # Loop through each grid square
        for i in range(grid_size[0]):  # Rows
            for j in range(grid_size[1]):  # Columns
                # Define the boundaries of the current grid square
                start_row = i * grid_height
                end_row = (i + 1) * grid_height
                start_col = j * grid_width
                end_col = (j + 1) * grid_width

                # Extract the part of the mask corresponding to the current grid square
                grid_square = predicted_mask[start_row:end_row, start_col:end_col]

                # Check if there are any crosswalk pixels (1's) in the grid square
                if np.any(grid_square == 1):
                    crosswalk_grid[i, j] = True

        return crosswalk_grid

    def calculate_accuracy_for_image(pred_intersection_matrix, actual_intersection_matrix, specific_correct_list, img_name):
        pred_accuracy = 0

        for x in range(4):
            for y in range(4):
                if pred_intersection_matrix[x][y] == actual_intersection_matrix[x][y]:
                    pred_accuracy += 1

        for num in specific_correct_list:
            if isinstance(num, str):
                if num.startswith(':'):
                    specific_number = num.split(':')[1]
                    if pred_accuracy <= int(specific_number):
                        specific_correct_dict[f'number_{num}'] += 1
            elif isinstance(num, int):
                if pred_accuracy == num:
                    specific_correct_dict[f'number_{num}'] += 1

        accuracy_for_img = pred_accuracy / 16

        in_image_accuracy_list.append(accuracy_for_img)
        if accuracy_for_img == 1:
            perfect_accuracy_list.append(1)
            perfect_accuracy_name_list.append(img_name[:-4])
        else:
            perfect_accuracy_list.append(0)

    def add_results_to_excel_spreadsheet(xlsx_file_path, specific_correct_list, step, model_name):
        in_image_accuracy = round((sum(in_image_accuracy_list) / len(in_image_accuracy_list) * 100), 2)
        perfect_accuracy = round((sum(perfect_accuracy_list) / len(perfect_accuracy_list) * 100), 2)

        wb_obj = openpyxl.load_workbook(xlsx_file_path)

        sheet = wb_obj.active

        max_row = get_max_row_with_data(sheet)
        max_column = get_max_column_with_data(sheet)

        bold_font = styles.Font(bold=True)
        standard_font = styles.Font(bold=False)

        # Check if the columns for the specific correct list values exist and creating them if needed
        for i, num in enumerate(specific_correct_list):
            add_num_to_col = True
            if isinstance(num, int):
                for column_value_index in range(1, max_column + 1):
                    if sheet.cell(row=1, column=column_value_index).value == f'sum {num}/16':
                        add_num_to_col = False
                        # if the specific correct values already have an assigned column, set the perfect_images_names two cols to the right
                        if i + 1 == len(specific_correct_list):
                            perfect_images_col = column_value_index + 1
                        break
            elif isinstance(num, str):
                formatted_num = num.split(':')[1]
                for column_value_index in range(1, max_column + 1):
                    if sheet.cell(row=1, column=column_value_index).value == f'{formatted_num} or below':
                        add_num_to_col = False
                        break

            if add_num_to_col:
                if isinstance(num, int):
                    sheet.cell(row=1, column=max_column + 1).value = f'sum {num}/16'
                    # If it's the last specific correct number, set the column for the perfect_images_names to 2 further to the right
                    if num == specific_correct_list[-1]:
                        perfect_images_col = max_column + 2
                elif isinstance(num, str):
                    sheet.cell(row=1, column=max_column + 1).value = f'{num.split(":")[1]} or below'
                sheet.cell(row=1, column=max_column + 1).font = bold_font
                wb_obj.save(os.path.join(os.getcwd(), xlsx_file_path))
                max_column = get_max_column_with_data(sheet)
                max_row = get_max_row_with_data(sheet)

        # Find the existing row for this exact model to add results
        model_name_exact = f'{model_name}_step{step}'
        for row_num in range(1, max_row + 1):
            model_name_of_sheet_line = sheet[f'A{row_num}'].value
            step_of_sheet_line = sheet[f'B{row_num}'].value
            warmup_steps_of_sheet_line = sheet[f'C{row_num}'].value

            if model_name_of_sheet_line == model_name_exact and step_of_sheet_line == step and warmup_steps_of_sheet_line == warmup_steps:
                active_row = row_num
                break

        # Add the specific correct numbers to the Excel spreadsheet (to the correct line)
        for num in specific_correct_list:
            if isinstance(num, int):
                for column_value_index in range(1, max_column + 1):
                    if sheet.cell(row=1, column=column_value_index).value == f'sum {num}/16':
                        sheet.cell(row=active_row, column=column_value_index).value = specific_correct_dict[f'number_{num}']
                        sheet.cell(row=active_row, column=column_value_index).font = standard_font

                        break
            elif isinstance(num, str):
                formatted_num = num.split(':')[1]
                for column_value_index in range(1, max_column + 1):
                    if sheet.cell(row=1, column=column_value_index).value == f'{formatted_num} or below':
                        sheet.cell(row=active_row, column=column_value_index).value = specific_correct_dict[f'number_{num}']
                        sheet.cell(row=active_row, column=column_value_index).font = standard_font
                        break

        # Add the names of the images the model predicted perfectly to the spreadsheet
        sheet.cell(row=active_row, column=perfect_images_col).value = f'{perfect_accuracy_name_list}'
        sheet.cell(row=active_row, column=perfect_images_col).font = standard_font

        # Add accuracy and perfect accuracy to spreadsheet
        sheet.cell(row=active_row, column=5).value = in_image_accuracy
        # sheet.cell(row=active_row, column=6).font = standard_font
        sheet.cell(row=active_row, column=6).value = perfect_accuracy
        # sheet.cell(row=active_row, column=7).font = standard_font

        wb_obj.save(os.path.join(os.getcwd(), xlsx_file_path))

    def open_actual_intersection_matrix(actual_matrix_dir, matrix):
        with open(os.path.join(actual_matrix_dir, matrix)) as actual_matrix_file:
            actual_matrix_content = actual_matrix_file.read()
            actual_matrix_content = ast.literal_eval(actual_matrix_content)
            return actual_matrix_content

    for step_dir in tqdm(os.listdir(ckpt_parent_dir)):
        in_image_accuracy_list = []
        perfect_accuracy_list = []
        perfect_accuracy_name_list = []
        specific_correct_dict = {}
        for number in specific_correct_list:
            specific_correct_dict[f'number_{number}'] = 0

        for mask_pred in os.listdir(os.path.join(ckpt_parent_dir, step_dir, 'predictions')):
            binary_mask = load_binary_mask(os.path.join(ckpt_parent_dir, step_dir, 'predictions', mask_pred))
            pred_intersection_matrix = check_crosswalk_in_grid(binary_mask)
            actual_intersection_matrix = open_actual_intersection_matrix(actual_matrix_dir, f'{mask_pred[:-4]}.txt')
            calculate_accuracy_for_image(pred_intersection_matrix, actual_intersection_matrix, specific_correct_list, mask_pred)
            add_results_to_excel_spreadsheet(xlsx_file_path, specific_correct_list, int(step_dir[4:]), model_name)






'''
Setting up Parameters required for running this script
'''
matplotlib.use('TkAgg')  # Or 'Qt5Agg', depending on your setup

train_steps = 100000
warmup_steps = 10000
model_version = 'v8'
model_name = 'DeepLabV3_MobilenetV2'
# STEP: TRAINING
train_data_tfrecords = os.path.join('data', 'tfrecords', 'train_bb.tfrecord')
val_data_tfrecords = os.path.join('data', 'tfrecords', 'val_trimap.tfrecord')
test_data_tfrecords = os.path.join('data', 'tfrecords', 'test_bb.tfrecord')
trained_model = os.path.join('models', model_name, model_version, 'trained_model')
# export_dir = os.path.join('models', model_name, model_version, 'checkpoints_steps', 'step200', 'saved_model')
# config model
model_ckpt_path = './model_ckpt/mnv2_deeplabv3_pascal/'
num_classes = 2  # foreground and background
WIDTH, HEIGHT = 120, 120
input_size = [HEIGHT, WIDTH, 3]
BATCH_SIZE = 4
train_img_dir = os.path.join('data', 'images', 'train_bb')
val_img_dir = os.path.join('data', 'images', 'val_bb')
test_img_dir = os.path.join('data', 'images', 'test_bb')
train_mask_dir = os.path.join('data/Trimaps/train_trimap')
val_mask_dir = os.path.join('data/Trimaps/val_trimap')
test_mask_dir = os.path.join('data/Trimaps/test_trimap')

amount_of_train_samples = len(os.listdir(train_img_dir))
amount_of_val_samples = len(os.listdir(val_img_dir))
amount_of_test_samples = len(os.listdir(test_img_dir))

model_dir = os.path.join(f'models/{model_name}/{model_version}/trained_model')

print(amount_of_train_samples, amount_of_val_samples, amount_of_test_samples)

# STEP: SAVING MODELS
ckpt_parent_dir = os.path.join('models', model_name, model_version, 'checkpoints_steps')
all_checkpoints = ["ckpt-495", "ckpt-496", "ckpt-497", "ckpt-498", "ckpt-499", "ckpt-500"]

# STEP: PREDICTING
plot_predictions = False
save_pred_binary_masks = True

# STEP: EVALUATING
specific_correct_list = [':4', 13, 14, 15]
actual_matrix_dir = os.path.join('..', 'train_val_test_split', 'captcha_matrices')
xlsx_file_path = os.path.join('excel_files', 'Segmentation_Optimization.xlsx')


'''
Executing desired Actions
'''
# supress warnings (they are cluttering the log --> remove if anything isn't working for further debugging)
absl.logging.set_verbosity(absl.logging.ERROR)

exp_config = config_model()
distribution_strategy, task = config_trainer()


train_model()
save_models()
print('predicting...')
get_predicted_masks_from_model()
print('evaluating models...')
evaluate_models()












